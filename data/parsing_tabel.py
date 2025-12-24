#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
detect_schedule.py

Применение:
    python detect_schedule.py

Файлы:
    читает data/График_ТМ5_февраль.xlsx
    записывает результаты в тот же файл, в первый пустой столбец справа:
        заголовок "Определённый график" (в строку 3),
        для каждой строки справа значение "4*2", "5*2" или "НЕ ОПРЕДЕЛЕН".
Вывод:
    в консоль — таб. номера сотрудников, для которых график не определился.
"""

from openpyxl import load_workbook
import re

FILE = "График_ТМ5_февраль.xlsx"
START_ROW = 4   # читать с 4-й строки (1-based)
START_COL = 5   # читать с 5-го столбца (1-based)

# параметры проверки
MATCH_THRESHOLD = 0.75   # минимальный процент совпадений среди известных ячеек
MIN_KNOWN = 5            # минимум известных ячеек для принятия решения

# шаблоны
PATTERN_4x2 = [1,1,1,1,0,0]        # период 6
PATTERN_5x2 = [1,1,1,1,1,0,0]      # период 7
PATTERNS = {
    '4*2': PATTERN_4x2,
    '5*2': PATTERN_5x2
}

# вспомогательные функции
def cell_to_flag(val):
    """Преобразование ячейки в флаг:
       1 - рабочий (1 или 2 содержащийся),
       0 - выходной (В или B),
       None - пропуск/О/непонятно
    """
    if val is None:
        return None
    s = str(val).strip()
    if s == '':
        return None
    # рукописные/русские В, латинские B
    if re.fullmatch(r'(?i)[вb]', s):
        return 0
    # буква О или o — пропускаем
    if re.fullmatch(r'(?i)o|о', s):
        return None
    # цифры 1 или 2 (возможно с точкой/запятой)
    if re.search(r'[12]', s):
        return 1
    # явный 0
    if re.fullmatch(r'0+(\.0+)?', s):
        return 0
    # остальные случаи — None
    return None

def best_pattern_match(obs_flags):
    """
    obs_flags: list из 1/0/None
    возвращает: (best_label, best_period, best_shift, match_ratio, known_count, predicted_sequence)
    или None, если ничего не подошло
    """
    n = len(obs_flags)
    best = None
    for label, pattern in PATTERNS.items():
        p = len(pattern)
        for shift in range(p):
            known = 0
            matches = 0
            predicted = []
            for i in range(n):
                pat_val = pattern[(i + shift) % p]   # циклически
                predicted.append(pat_val)
                if obs_flags[i] is None:
                    continue
                known += 1
                if obs_flags[i] == pat_val:
                    matches += 1
            if known == 0:
                continue
            ratio = matches / known
            # сохраняем лучший
            if best is None or ratio > best[3]:
                best = (label, p, shift, ratio, known, predicted)
    if best is None:
        return None
    # условие принятия
    label, p, shift, ratio, known, predicted = best
    if ratio >= MATCH_THRESHOLD and known >= MIN_KNOWN:
        return best
    return None

# --- основной процесс ---
wb = load_workbook(FILE)
ws = wb.active

max_col = ws.max_column
max_row = ws.max_row

# 1) удалить полностью пустые строки (сверху вниз — лучше идти снизу вверх при удалении)
# считаем строку пустой, если все ячейки в диапазоне 1..max_col пусты/пустые строки
rows_deleted = 0
for r in range(max_row, START_ROW - 1, -1):
    empty = True
    for c in range(1, max_col + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != '':
            empty = False
            break
    if empty:
        ws.delete_rows(r, 1)
        rows_deleted += 1

if rows_deleted > 0:
    print(f"Удалено пустых строк: {rows_deleted}")

# обновим лимиты
max_col = ws.max_column
max_row = ws.max_row

# найдем столбец для записи результата (первый пустой справа)
result_col = max_col + 1

# запишем заголовок (предполагаю, что заголовки находятся в строке 3 — если нет, можно изменить)
HEADER_ROW = 3
ws.cell(row=HEADER_ROW, column=result_col).value = "Определённый график"

unknown_tabs = []

# для каждой строки с START_ROW до конца
for r in range(START_ROW, max_row + 1):
    # получим таб. номер из первого столбца (A)
    tab = ws.cell(row=r, column=1).value
    # прочитаем последовательность справа от START_COL до конца (или до тех пор, пока не встретится полностью пустая серия)
    obs = []
    for c in range(START_COL, max_col + 1):
        v = ws.cell(row=r, column=c).value
        flag = cell_to_flag(v)
        obs.append(flag)

    # если в строке вообще нет данных в рассматриваемой части, пропустим
    if all(x is None for x in obs):
        ws.cell(row=r, column=result_col).value = "ПУСТО"
        continue

    # попытаемся найти лучший шаблон (сдвиги)
    res = best_pattern_match(obs)
    if res is not None:
        label, p, shift, ratio, known, predicted = res
        ws.cell(row=r, column=result_col).value = label
    else:
        # fallback: попробуем более простое правило — найти максимальное подряд идущих рабочих (между В) и если
        # видим блок длиной 4 или 5, принять его (учтёт случаи когда месяц начался в середине блока)
        # преобразуем obs к строке символов, пропуская None и О, оставляя В и 1
        simple = []
        for x in obs:
            if x is None:
                simple.append('?')
            elif x == 1:
                simple.append('1')
            else:
                simple.append('V')  # V означает В (выходной)
        # объединяем подрядные рабочие (учитываем цикличность - если начало и конец рабочие, то возможно объединить)
        # найдём максимальную длину блока рабочих (не считая '?')
        max_block = 0
        current = 0
        for ch in simple:
            if ch == '1':
                current += 1
                if current > max_block:
                    max_block = current
            elif ch == '?':
                current = 0
            else:
                current = 0
        # циклический случай: если начало и конец оба '1', соединим
        if simple and simple[0] == '1' and simple[-1] == '1':
            # считаем подряд с конца
            end_block = 0
            for ch in reversed(simple):
                if ch == '1':
                    end_block += 1
                else:
                    break
            start_block = 0
            for ch in simple:
                if ch == '1':
                    start_block += 1
                else:
                    break
            if end_block + start_block > max_block:
                max_block = end_block + start_block

        if max_block >= 5:
            ws.cell(row=r, column=result_col).value = "5*2"
        elif max_block == 4:
            ws.cell(row=r, column=result_col).value = "4*2"
        else:
            ws.cell(row=r, column=result_col).value = "НЕ ОПРЕДЕЛЕН"
            unknown_tabs.append(tab if tab is not None else f"строка {r}")

# сохраним файл
wb.save(FILE)
print(f"Результат записан в файл: {FILE}")

# выведем неизвестные таб. номера
if unknown_tabs:
    print("Не удалось определить график для следующих таб. номеров:")
    for t in unknown_tabs:
        print("  ", t)
else:
    print("Все графики определены.")
