from openpyxl import load_workbook, Workbook
import re

INPUT_FILE = "График_ТМ5_февраль.xlsx"
OUTPUT_FILE = "График_ТМ5_февраль_ГОТОВО.xlsx"

# номера столбцов (1-based)
COL_TAB = 1
COL_SMENA = 3
COL_VYH = 4
COL_DAYS_START = 5
COL_DAYS_END = 35
COL_GRAPH = 38

START_ROW = 4  # данные с 4 строки

wb_in = load_workbook(INPUT_FILE)
ws_in = wb_in.active

wb_out = Workbook()
ws_out = wb_out.active

# --- заголовки ---
headers = (
    ["Таб.№", "График", "Режим", "см.", "вых."] +
    [str(i) for i in range(1, 32)]
)

for col, header in enumerate(headers, start=1):
    ws_out.cell(row=1, column=col).value = header

out_row = 2

# --- обработка строк ---
for r in range(START_ROW, ws_in.max_row + 1):
    tab = ws_in.cell(r, COL_TAB).value
    if tab is None or str(tab).strip() == "":
        continue  # пропуск пустых строк

    # --- график ---
    graph = ws_in.cell(r, COL_GRAPH).value
    if graph is not None:
        graph = str(graph).replace("*", "х")

    # --- режим ---
    has_1 = False
    has_2 = False

    days_values = []

    for c in range(COL_DAYS_START, COL_DAYS_END + 1):
        val = ws_in.cell(r, c).value
        if val is None:
            days_values.append("")
            continue

        s = str(val).strip()

        if s == "1":
            has_1 = True
        elif s == "2":
            has_2 = True

        days_values.append(s)

    if has_1 and has_2:
        regime = "1,2"
    elif has_1:
        regime = "1"
    elif has_2:
        regime = "2"
    else:
        regime = "—"

    # --- запись строки ---
    ws_out.cell(out_row, 1).value = tab
    ws_out.cell(out_row, 2).value = graph
    ws_out.cell(out_row, 3).value = regime
    ws_out.cell(out_row, 4).value = ws_in.cell(r, COL_SMENA).value
    ws_out.cell(out_row, 5).value = ws_in.cell(r, COL_VYH).value

    for i, day_val in enumerate(days_values):
        ws_out.cell(out_row, 6 + i).value = day_val

    out_row += 1

wb_out.save(OUTPUT_FILE)

print("Готово.")
print(f"Создан файл: {OUTPUT_FILE}")
