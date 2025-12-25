from datetime import datetime, timedelta
from openpyxl import load_workbook
import json
from pathlib import Path

from structure_model.config import (
    BASE_DIR,
    FILE_PATH,
    ROW_START,
    STEP,
    COL_SHIFT_1_START,
    COL_SHIFT_1_END,
    COL_SHIFT_2_START,
    COL_SHIFT_2_END,
    COL_SHIFT_1_INSERT,
    COL_SHIFT_2_INSERT,
    REST_HOURS,
    ALLOW_WEEKEND_EXTRA_WORK,
    ABSENCES_FILE,
    TOTAL_DAYS_IN_MONTH,
    TRANSPORTS,
)

from structure_model.excel_io import get_schedule_slots
from structure_model.history_manager import load_history, save_history


# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =============================================================================

def load_absent_drivers(day: int, shift: int) -> set[str]:
    if not ABSENCES_FILE or not Path(ABSENCES_FILE).exists():
        return set()

    try:
        with open(ABSENCES_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return set()

    absent = set()
    for rec in data:
        try:
            if int(rec["day"]) == day and int(rec["shift"]) == shift:
                absent.add(str(rec["tab_no"]).strip())
        except Exception:
            continue

    return absent


def get_rest_hours(driver_id, history, target_start):
    drv = str(driver_id)
    if drv not in history:
        return REST_HOURS

    last = history[drv]
    try:
        h, m = map(int, last["end_str"].split(":"))
        base_date = target_start.date() - timedelta(days=1)
        if last.get("is_next_day"):
            base_date += timedelta(days=1)

        last_end = datetime.combine(base_date, datetime.min.time()).replace(hour=h, minute=m)
        return (target_start - last_end).total_seconds() / 3600
    except Exception:
        return -9999


def worked_same_shift_yesterday(driver_id, history, shift):
    drv = str(driver_id)
    return drv in history and history[drv].get("shift_code") == shift


def choose_driver(candidates, history, shift_start, shift, assigned_today):
    scored = []

    for drv in candidates:
        if drv in assigned_today:
            continue

        rest = get_rest_hours(drv, history, shift_start)
        if rest < 0:
            continue

        same_shift_penalty = 1 if worked_same_shift_yesterday(drv, history, shift) else 0
        scored.append((abs(rest - REST_HOURS), same_shift_penalty, -rest, drv))

    if not scored:
        return None

    scored.sort()
    return scored[0][3]


# =============================================================================
# ОСНОВНОЙ ПЛАНИРОВЩИК
# =============================================================================

def run_planner(day: int, prev_day: int, transport: str, route: str):
    print("\n" + "=" * 80)
    print(f"[RUN] day={day}, transport={transport}, route={route}")

    now = datetime.now()
    try:
        target_date = datetime(now.year, now.month, day)
    except ValueError:
        print("[ERROR] Некорректный день")
        return

    is_weekend = target_date.weekday() >= 5

    cfg = TRANSPORTS[transport]
    sheets = cfg["sheets"]
    output_root = cfg["output_dir"]

    sheet_name = sheets.get((route, is_weekend))
    if not sheet_name:
        print(f"[SKIP] Нет листа (route={route}, weekend={is_weekend})")
        return

    print(f"[INFO] Excel-лист: {sheet_name}")

    history = load_history(prev_day)
    today_history = {}

    # --- slots ---
    slots_s1 = get_schedule_slots(
        FILE_PATH, ROW_START, STEP,
        COL_SHIFT_1_START, COL_SHIFT_1_END, 1, sheet_name
    )
    slots_s2 = get_schedule_slots(
        FILE_PATH, ROW_START, STEP,
        COL_SHIFT_2_START, COL_SHIFT_2_END, 2, sheet_name
    )

    # --- consolidation ---
    cons_path = BASE_DIR / "consolidation" / transport / route / "data.json"
    if not cons_path.exists():
        print(f"[ERROR] Нет consolidation: {cons_path}")
        return

    with open(cons_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    drivers = {
        str(e["tab_number"]).strip()
        for e in data.get("employees", [])
        if e.get("tab_number") is not None
    }

    if not drivers:
        print("[ERROR] Пустой список водителей")
        return

    print(f"[INFO] Водителей: {len(drivers)}")

    # --- absences ---
    absent_s1 = load_absent_drivers(day, 1)
    absent_s2 = load_absent_drivers(day, 2)

    cand_s1 = drivers - absent_s1
    cand_s2 = drivers - absent_s2

    # --- output ---
    route_dir = output_root / route
    route_dir.mkdir(parents=True, exist_ok=True)
    out_file = route_dir / f"Расписание_Итог_{day}.xlsx"

    wb = load_workbook(FILE_PATH)
    for sh in wb.sheetnames[:]:
        if sh != sheet_name:
            del wb[sh]

    ws = wb[sheet_name]
    ws.title = "Расписание"

    for s in slots_s1:
        ws.cell(s["excel_row"], COL_SHIFT_1_INSERT).value = None
    for s in slots_s2:
        ws.cell(s["excel_row"], COL_SHIFT_2_INSERT).value = None

    assigned_today = set()

    # --- shift 1 ---
    for s in slots_s1:
        drv = choose_driver(cand_s1, history, s["time_info"]["start_dt"], 1, assigned_today)
        if not drv and ALLOW_WEEKEND_EXTRA_WORK:
            drv = choose_driver(drivers, history, s["time_info"]["start_dt"], 1, assigned_today)

        ws.cell(s["excel_row"], COL_SHIFT_1_INSERT).value = drv or "НЕТ_РЕЗЕРВА"
        if drv:
            assigned_today.add(drv)
            cand_s1.discard(drv)
            cand_s2.discard(drv)
            today_history[drv] = {**s["time_info"], "shift_code": 1}

    # --- shift 2 ---
    for s in slots_s2:
        drv = choose_driver(cand_s2, history, s["time_info"]["start_dt"], 2, assigned_today)
        if not drv and ALLOW_WEEKEND_EXTRA_WORK:
            drv = choose_driver(drivers, history, s["time_info"]["start_dt"], 2, assigned_today)

        ws.cell(s["excel_row"], COL_SHIFT_2_INSERT).value = drv or "НЕТ_РЕЗЕРВА"
        if drv:
            assigned_today.add(drv)
            cand_s2.discard(drv)
            cand_s1.discard(drv)
            today_history[drv] = {**s["time_info"], "shift_code": 2}

    wb.save(out_file)
    save_history(day, today_history)

    print(f"[DONE] {out_file} | назначено: {len(today_history)}")


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    for transport, cfg in TRANSPORTS.items():
        for day_type, routes in cfg["routes"].items():
            print(f"[INFO] {transport} / {day_type}: {routes}")

        for day in range(1, TOTAL_DAYS_IN_MONTH + 1):
            prev = max(day - 1, 0)
            is_weekend = datetime.now().replace(day=day).weekday() >= 5
            route_list = cfg["routes"]["weekend" if is_weekend else "workday"]

            for route in route_list:
                run_planner(day, prev, transport, route)
