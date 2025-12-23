import pandas as pd
import json
import os
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta

# --- Конфигурация папок и файлов ---
HISTORY_JSON_DIR = "../history_json"
OUTPUT_DIR = "../output"

os.makedirs(HISTORY_JSON_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

FILE_PATH = "../data.xlsx"  # твой excel-файл
TARGET_DAY = 1
PREV_DAY = 0
TOTAL_DAYS_IN_MONTH = 30

# --- Настройки чтения/записи Excel (по твоей разметке) ---
ROW_START = 5  # Начальная строка с данными (5-я строка Excel)
STEP = 2  # Шаг строки (через одну)
COL_SHIFT_1_INSERT = 3  # Куда писать 1 смену (вставка)
COL_SHIFT_2_INSERT = 8  # Куда писать 2 смену (вставка)

COL_SHIFT_1_START = 5  # Где читать начало 1 смены
COL_SHIFT_1_END = 6  # Где читать окончание 1 смены

COL_SHIFT_2_START = 10  # Где читать начало 2 смены
COL_SHIFT_2_END = 11  # Где читать окончание 2 смены

# --- Политика отдыха ---
# REST_POLICY: "double" = отдых = 2 * длительность работы (как раньше)
#              "min12"  = отдых = max(2 * длительность, 12) (не менее 12 часов)
REST_POLICY = "min12"   # "double" или "min12" — по твоему требованию

# Если True — при отсутствии валидных кандидатов разрешаем взять
# того, кто почти отдохнул (ожидание <= WAIT_THRESHOLD_HOURS).
# По умолчанию отключено — безопаснее.
ALLOW_ASSIGN_BEFORE_REST = False
WAIT_THRESHOLD_HOURS = 3.0

# --- Параметры гибкого старта и ограничений работы ---
MIN_WORK_HOURS = 6.0    # минимум часов работы по норме
MAX_WORK_HOURS = 10.0   # максимум часов работы по норме
# Максимально допустимое смещение старта в пределах слота (чтобы не выходить за разум)
# На практике мы считаем допустимым смещением начало на любую величину внутри слота,
# но стартап должен позволить отработать хотя бы MIN_WORK_HOURS.
# (поэтому фактический лимит задаётся проверкой на overlap)
# -------------------------

def get_shift_info_from_two_cols(start_val, end_val):
    """
    Парсит две ячейки и возвращает словарь:
    start_dt, end_dt (datetime на текущую дату времени now),
    start_str, end_str, duration (часы), is_next_day
    Если не удалось — возвращает None.
    """
    if pd.isna(start_val) or pd.isna(end_val):
        return None

    try:
        start_str = str(start_val).strip().split()[0]
        end_str = str(end_val).strip().split()[0]

        match_start = re.search(r'(\d{1,2})[:\.\-](\d{2})', start_str)
        match_end = re.search(r'(\d{1,2})[:\.\-](\d{2})', end_str)

        if not match_start or not match_end:
            return None

        h1, m1 = map(int, match_start.groups())
        h2, m2 = map(int, match_end.groups())

        now = datetime.now().replace(second=0, microsecond=0)
        start_dt = now.replace(hour=h1, minute=m1)
        end_dt = now.replace(hour=h2, minute=m2)

        is_next_day = False
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
            is_next_day = True

        duration = (end_dt - start_dt).total_seconds() / 3600

        return {
            'start_dt': start_dt,
            'end_dt': end_dt,
            'start_str': f"{h1:02}:{m1:02}",
            'end_str': f"{h2:02}:{m2:02}",
            'duration': round(duration, 2),
            'is_next_day': is_next_day
        }
    except Exception:
        return None


def calculate_rest_duration(end_str_N, is_next_day_N, start_str_N1):
    """
    Вычисляет фактическое количество часов отдыха между окончанием N-ной смены
    и началом следующей смены N+1, ориентируясь на текущие даты.
    (Используется при генерации отчёта).
    """
    try:
        h_end, m_end = map(int, end_str_N.split(':'))
        h_start, m_start = map(int, start_str_N1.split(':'))
        base_dt = datetime.now().replace(second=0, microsecond=0)
        end_dt_N = base_dt.replace(hour=h_end, minute=m_end)
        if not is_next_day_N:
            end_dt_N -= timedelta(days=1)
        start_dt_N1 = base_dt.replace(hour=h_start, minute=m_start)
        while start_dt_N1 <= end_dt_N:
            start_dt_N1 += timedelta(days=1)
        rest_duration = (start_dt_N1 - end_dt_N).total_seconds() / 3600
        return round(rest_duration, 1)
    except Exception:
        return "Н/Д"


# --- Работа с историей (чтение/запись JSON) ---

def load_history(day_num):
    filename = os.path.join(HISTORY_JSON_DIR, f"history_{day_num}.json")
    if not os.path.exists(filename):
        return {}
    with open(filename, 'r', encoding='utf-8') as f:
        try:
            return json.load(f)
        except Exception:
            return {}


def save_history(day_num, data):
    filename = os.path.join(HISTORY_JSON_DIR, f"history_{day_num}.json")
    clean_data = {}
    for k, v in data.items():
        clean_data[k] = {}
        if isinstance(v, dict):
            for key, val in v.items():
                if isinstance(val, datetime):
                    clean_data[k][key] = val.isoformat()
                else:
                    clean_data[k][key] = val
        else:
            clean_data[k] = v
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(clean_data, f, indent=4, ensure_ascii=False)
    print(f"[Память] Данные о сменах сохранены в {filename}")


# -------------------------
# Восстановление окончания последней смены
# -------------------------

def _reconstruct_last_end_dt(end_str, is_next_day, target_start_dt):
    try:
        h, m = map(int, end_str.split(':'))
    except Exception:
        return target_start_dt - timedelta(hours=24)

    c_same = datetime(target_start_dt.year, target_start_dt.month, target_start_dt.day, h, m)
    c_prev = c_same - timedelta(days=1)
    c_next = c_same + timedelta(days=1)
    candidates = [c_same, c_prev, c_next]
    valid = [c for c in candidates if c <= target_start_dt]
    if valid:
        return max(valid)
    return min(candidates, key=lambda x: abs((target_start_dt - x).total_seconds()))


# -------------------------
# Правила отдыха + гибкий старт
# -------------------------

def filter_by_rest_and_flexible_start(drivers, history_data, slot_start_dt, slot_end_dt):
    """
    Для каждого водителя вычисляем:
      - last_end_dt (восстановленный)
      - needed_rest_hours по REST_POLICY
      - earliest_start_dt = last_end_dt + needed_rest_hours
    Возвращаем:
      valid: список водителей, которые могут стартовать не позже slot_start_dt (т.е. уже отдохнули)
      flexible: список кортежей (drv, earliest_start_dt, wait_h), где earliest_start_dt > slot_start_dt
                но есть возможность начать позже внутри слота и при этом отработать >= MIN_WORK_HOURS
      banned: те, у кого полностью не хватает времени для MIN_WORK_HOURS в этом слоте
    """
    valid = []
    flexible = []  # (drv, earliest_start_dt, wait_h)
    banned = []     # (drv, wait_h)

    for drv in drivers:
        drv_str = str(drv)
        if drv_str not in history_data:
            valid.append((drv, None))  # None означает нет ограничений по истории
            continue

        last_work = history_data.get(drv_str, {})
        try:
            last_duration = float(last_work.get('duration', 0))
        except Exception:
            last_duration = 0.0

        last_end_str = last_work.get('end_str')
        was_next_day = last_work.get('is_next_day', False)

        if not last_end_str:
            valid.append((drv, None))
            continue

        last_end_dt = _reconstruct_last_end_dt(last_end_str, was_next_day, slot_start_dt)

        if REST_POLICY == "double":
            needed_rest_hours = last_duration * 2.0
        else:  # "min12"
            needed_rest_hours = max(last_duration * 2.0, 12.0)

        earliest_start_dt = last_end_dt + timedelta(hours=needed_rest_hours)

        if earliest_start_dt <= slot_start_dt:
            # водитель полностью отдохнул до начала слота
            valid.append((drv, earliest_start_dt))
            continue

        # если earliest_start_dt > slot_start_dt — проверим можно ли начать позже в пределах слота
        # и при этом отработать хотя бы MIN_WORK_HOURS (и не более MAX_WORK_HOURS)
        # допустимая рабочая зона: [max(slot_start_dt, earliest_start_dt), slot_end_dt]
        possible_start = earliest_start_dt
        # если earliest_start_dt < slot_start_dt (handled above), else >=
        if possible_start < slot_start_dt:
            possible_start = slot_start_dt

        # Максимальный возмож рабочий интервал если стартуем в possible_start:
        max_possible_duration = (slot_end_dt - possible_start).total_seconds() / 3600.0
        if max_possible_duration + 1e-6 >= MIN_WORK_HOURS:
            # можно назначить гибко: ограничим duration в границах [MIN_WORK_HOURS, MAX_WORK_HOURS, max_possible_duration]
            flexible.append((drv, earliest_start_dt, round((earliest_start_dt - slot_start_dt).total_seconds() / 3600.0, 1)))
        else:
            # не хватает времени внутри слота, баним на этот слот
            wait_h = (earliest_start_dt - slot_start_dt).total_seconds() / 3600.0
            banned.append((drv, round(wait_h, 1)))

    # сортируем
    valid.sort(key=lambda x: (0 if x[1] is None else 1, x[1] if x[1] else datetime.min))
    flexible.sort(key=lambda x: x[2])  # минимальное ожидание первым
    banned.sort(key=lambda x: x[1])

    # Возвращаем в более удобной форме
    valid_ids = [v[0] for v in valid]
    return valid_ids, flexible, banned


# -------------------------
# Чтение расписания и табеля
# -------------------------

def get_schedule_slots(file_path, start_row, step, col_start, col_end, shift_code):
    df = pd.read_excel(file_path, sheet_name="Расписание", header=None)
    slots = []
    current_row = start_row

    print(f"--- АНАЛИЗ {shift_code} СМЕНЫ (Колонки {col_start}/{col_end}) ---")

    while True:
        pd_idx = current_row - 1
        if pd_idx >= len(df):
            break

        start_val = df.iloc[pd_idx, col_start]
        end_val = df.iloc[pd_idx, col_end]

        time_info = get_shift_info_from_two_cols(start_val, end_val)

        if time_info:
            time_info['shift_code'] = shift_code
            slots.append({
                'excel_row': current_row,
                'time_info': time_info
            })
            print(f"ОК: Строка {current_row:<3}: {time_info['start_str']} - {time_info['end_str']} ({time_info['duration']} ч.)")
        current_row += step

    return slots


def get_available_drivers(file_path, day_num, shift_code):
    df = pd.read_excel(file_path, sheet_name="Табель")
    df.columns = df.columns.astype(str)

    day_col = str(day_num)
    if day_col not in df.columns:
        if f'{day_num}.0' in df.columns:
            day_col = f'{day_num}.0'
        else:
            raise ValueError(f"Ошибка: В табеле нет колонки с названием '{day_num}'")

    drivers = []
    for idx, row in df.iterrows():
        tab_no = row.iloc[0]
        status = str(row[day_col]).strip()

        if pd.notna(tab_no) and status == str(shift_code):
            drivers.append(tab_no)

    return drivers


# -------------------------
# Планировщик (с гибким стартом)
# -------------------------

def run_planner():
    print(f"=== ЗАПУСК ПЛАНИРОВЩИКА: ДЕНЬ {TARGET_DAY} (С учетом истории дня {PREV_DAY}) ===\n")

    history = load_history(PREV_DAY)
    today_history_log = {}

    slots_s1 = get_schedule_slots(FILE_PATH, ROW_START, STEP, COL_SHIFT_1_START, COL_SHIFT_1_END, 1)
    slots_s2 = get_schedule_slots(FILE_PATH, ROW_START, STEP, COL_SHIFT_2_START, COL_SHIFT_2_END, 2)

    print(f"[Спрос] 1 смена: {len(slots_s1)} нарядов, 2 смена: {len(slots_s2)} нарядов")

    try:
        candidates_s1 = get_available_drivers(FILE_PATH, TARGET_DAY, 1)
        candidates_s2 = get_available_drivers(FILE_PATH, TARGET_DAY, 2)
    except ValueError as e:
        print(e)
        return

    print(f"[Табель] Доступно: 1 смена = {len(candidates_s1)}, 2 смена = {len(candidates_s2)}")

    wb = load_workbook(FILE_PATH)
    ws = wb["Расписание"]

    assigned_s1 = []
    assigned_s2 = []

    def process_slot(slot, candidates, insert_col, shift_code):
        nonlocal history, today_history_log, ws, assigned_s1, assigned_s2
        slot_start = slot['time_info']['start_dt']
        slot_end = slot['time_info']['end_dt']
        slot_duration = slot['time_info']['duration']

        # комбинированная история: предыдущие дни + уже назначенные сегодня
        combined_history = {**history, **today_history_log}

        valid, flexible, banned = filter_by_rest_and_flexible_start(candidates, combined_history, slot_start, slot_end)

        chosen = None
        chosen_start_dt = None
        chosen_end_dt = None
        chosen_duration = None

        # 1) если есть valid — выбираем по предпочтению (тот же shift_code в истории)
        if valid:
            prefer = [d for d in valid if history.get(str(d), {}).get('shift_code') == shift_code]
            if prefer:
                chosen = prefer[0]
            else:
                # предпочтение — тот, кто дольше всех отдыхал (есть в combined_history)
                def last_end_for(d):
                    rec = combined_history.get(str(d), {})
                    es = rec.get('end_str')
                    if not es:
                        return datetime.min
                    return _reconstruct_last_end_dt(rec.get('end_str'), rec.get('is_next_day', False), slot_start)
                valid_sorted = sorted(valid, key=lambda x: last_end_for(x))
                chosen = valid_sorted[0]

            # назначаем на обычный слот (старт = slot_start, end = slot_end or truncated by max work)
            # стараемся держать duration в пределах MIN..MAX (но не более slot duration)
            assigned_possible_duration = min(slot_duration, MAX_WORK_HOURS)
            if assigned_possible_duration < MIN_WORK_HOURS:
                # если слот короче минимума — не назначаем
                chosen = None
            else:
                chosen_start_dt = slot_start
                # end prefer: start + assigned_possible_duration but not after slot_end
                chosen_end_dt = min(slot_start + timedelta(hours=assigned_possible_duration), slot_end)
                chosen_duration = round((chosen_end_dt - chosen_start_dt).total_seconds() / 3600.0, 2)

        # 2) если нет valid, но есть flexible — попробуем первого flex (минимальный wait)
        if not chosen and flexible:
            # flexible contains tuples (drv, earliest_start_dt, wait_h)
            drv, earliest_start_dt, wait_h = flexible[0]
            # возможный старт = earliest_start_dt
            # максим возможный рабочий интервал:
            max_dur = (slot_end - earliest_start_dt).total_seconds() / 3600.0
            # выбираем duration = clamp between MIN_WORK_HOURS and min(MAX_WORK_HOURS, max_dur)
            dur = min(MAX_WORK_HOURS, max_dur)
            if dur + 1e-6 >= MIN_WORK_HOURS:
                chosen = drv
                chosen_start_dt = earliest_start_dt
                chosen_end_dt = earliest_start_dt + timedelta(hours=min(dur, MAX_WORK_HOURS))
                # но не выходить за slot_end
                if chosen_end_dt > slot_end:
                    chosen_end_dt = slot_end
                chosen_duration = round((chosen_end_dt - chosen_start_dt).total_seconds() / 3600.0, 2)

        # 3) если всё ещё нет chosen — попробуем взять из banned при ALLOW_ASSIGN_BEFORE_REST (малый wait)
        if not chosen and ALLOW_ASSIGN_BEFORE_REST and banned:
            candidate, wait_h = banned[0]
            if wait_h <= WAIT_THRESHOLD_HOURS:
                # если ожидание небольшое — назначаем обычный старт, но это уменьшит отдых чуть-чуть
                # следим за duration в лимитах
                if slot_duration >= MIN_WORK_HOURS:
                    chosen = candidate
                    chosen_start_dt = slot_start
                    chosen_end_dt = min(slot_start + timedelta(hours=min(slot_duration, MAX_WORK_HOURS)), slot_end)
                    chosen_duration = round((chosen_end_dt - chosen_start_dt).total_seconds() / 3600.0, 2)

        # если ничего не получилось — НЕТ_РЕЗЕРВА
        if not chosen:
            ws.cell(row=slot['excel_row'], column=insert_col).value = "НЕТ_РЕЗЕРВА"
            return None

        # final: записываем выбранного и его фактический start/end/duration
        # форматируем строки HH:MM
        s_str = f"{chosen_start_dt.hour:02}:{chosen_start_dt.minute:02}"
        e_str = f"{chosen_end_dt.hour:02}:{chosen_end_dt.minute:02}"
        is_next = chosen_end_dt.date() != chosen_start_dt.date()

        ws.cell(row=slot['excel_row'], column=insert_col).value = chosen

        # кладём в today_history_log: start_str, end_str, duration, is_next_day, shift_code
        today_history_log[str(chosen)] = {
            'start_str': s_str,
            'end_str': e_str,
            'duration': chosen_duration,
            'is_next_day': is_next,
            'shift_code': shift_code
        }

        # убрать выбранного из списка кандидатов (чтобы не назначить дважды)
        try:
            candidates.remove(chosen)
        except Exception:
            pass

        # track
        if shift_code == 1:
            assigned_s1.append(chosen)
        else:
            assigned_s2.append(chosen)

        return chosen

    # Назначаем обе смены
    for slot in slots_s1:
        process_slot(slot, candidates_s1, COL_SHIFT_1_INSERT, 1)

    for slot in slots_s2:
        process_slot(slot, candidates_s2, COL_SHIFT_2_INSERT, 2)

    print(f"\n[Итог] Назначено: 1 смена = {len(assigned_s1)}, 2 смена = {len(assigned_s2)}")
    file_path = os.path.join(OUTPUT_DIR, f"Расписание_Итог_{TARGET_DAY}.xlsx")
    wb.save(file_path)
    print(f"[Excel] Результат сохранен в '{file_path}'")

    save_history(TARGET_DAY, today_history_log)


# -------------------------
# Генерация сводного отчёта
# -------------------------

def generate_work_summary(target_day, file_path):
    print("\n\n=== ГЕНЕРАЦИЯ СВОДНОГО ОТЧЕТА ===")

    try:
        df_tab = pd.read_excel(file_path, sheet_name="Табель")
        driver_id_col = df_tab.columns[0]
        graphik_col = df_tab.columns[1]

        df_temp = df_tab.set_index(df_tab[driver_id_col].astype(str))

        driver_graphik = df_temp[graphik_col].to_dict()
        all_drivers = [str(d) for d in df_tab[driver_id_col].dropna().tolist()]
    except Exception as e:
        print(f"Ошибка чтения Табеля для отчета: {e}")
        return None

    history_data = {}
    for day in range(1, target_day + 2):
        filename = os.path.join(HISTORY_JSON_DIR, f"history_{day}.json")
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                try:
                    history_data[day] = json.load(f)
                except Exception:
                    history_data[day] = {}
        else:
            history_data[day] = {}

    columns = ['График'] + [f'День {d}' for d in range(1, target_day + 1)]
    report_data = []

    for driver_id in all_drivers:
        row = {'График': driver_graphik.get(driver_id, 'N/A')}

        for day in range(1, target_day + 1):
            day_history = history_data.get(day, {})

            if driver_id in day_history:
                work_data = day_history[driver_id]
                duration_N = work_data.get('duration', 0)
                end_str_N = work_data.get('end_str', '')
                is_next_day_N = work_data.get('is_next_day', False)

                next_day_history = history_data.get(day + 1, {})

                if driver_id in next_day_history:
                    start_str_N1 = next_day_history[driver_id].get('start_str', '')
                    rest_h = calculate_rest_duration(end_str_N, is_next_day_N, start_str_N1)
                    rest_report = f"Отдых: {rest_h} ч."
                elif day == target_day:
                    rest_report = "Отдых: Н/Д"
                else:
                    rest_report = "Отдых: Полный"

                row[f'День {day}'] = f"Работа: {duration_N} ч. | {rest_report}"
            else:
                row[f'День {day}'] = "--- Выходной"

        report_data.append(row)

    df_report = pd.DataFrame(report_data, index=pd.Index(all_drivers, name='Таб. №'))

    report_filename = os.path.join(OUTPUT_DIR, f"Отчет_Нагрузки_Дни_1_по_{target_day}.xlsx")
    df_report.to_excel(report_filename)
    print(f"\n[Отчет] Создан сводный отчет: **{report_filename}**")

    return report_filename


# -------------------------
# Симуляция на несколько дней (удаляет старые логи)
# -------------------------

def auto_run_simulation(total_days, file_path):
    print(f"\n--- ПОДГОТОВКА: Удаление старых log-файлов (history_X.json) ---\n")
    for d in range(1, total_days + 2):
        filename = os.path.join(HISTORY_JSON_DIR, f"history_{d}.json")
        if os.path.exists(filename):
            os.remove(filename)

    for day in range(1, total_days + 1):
        global TARGET_DAY
        global PREV_DAY

        TARGET_DAY = day
        PREV_DAY = day - 1

        print("\n" + "=" * 80)
        print(f"============================ ЗАПУСК ДНЯ {TARGET_DAY:02d} ============================")
        print("=" * 80)

        run_planner()

    print("\n" + "#" * 80)
    print("##################### СИМУЛЯЦИЯ ЗАВЕРШЕНА #####################")
    print("#" * 80)
    generate_work_summary(total_days, file_path)


if __name__ == "__main__":
    auto_run_simulation(TOTAL_DAYS_IN_MONTH, FILE_PATH)
