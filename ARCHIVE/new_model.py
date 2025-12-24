import pandas as pd
import json
import os
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Параметр маршрута - можно изменить при запуске
ROUTE_NUMBER = 55  # По умолчанию маршрут 55

# Пути зависят от маршрута
HISTORY_JSON_DIR = f"history_json_route_{ROUTE_NUMBER}"
OUTPUT_DIR = f"output_route_{ROUTE_NUMBER}"

os.makedirs(HISTORY_JSON_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Путь к файлу данных - можно использовать data_55.xlsx, data_9.xlsx и т.д.
# Если файл с номером маршрута не найден, используется data.xlsx
FILE_PATH = f"data_{ROUTE_NUMBER}.xlsx" if os.path.exists(f"data_{ROUTE_NUMBER}.xlsx") else "data.xlsx"
TARGET_DAY = 1
PREV_DAY = 0
TOTAL_DAYS_IN_MONTH = 30

# Параметры Excel
ROW_START = 5  # Начальная строка с данными (5-я строка Excel)
STEP = 2  # Шаг строки (через одну)
COL_SHIFT_1_INSERT = 3  # Куда писать 1 смену
COL_SHIFT_2_INSERT = 8  # Куда писать 2 смену

COL_SHIFT_1_START = 5  # Где читать начало 1 смены
COL_SHIFT_1_END = 6  # Где читать окончание 1 смены

COL_SHIFT_2_START = 10  # Где читать начало 2 смены
COL_SHIFT_2_END = 11  # Где читать окончание 2 смены

# Новая глобальная норма отдыха
REST_HOURS = 12.0


def get_shift_info_from_two_cols(start_val, end_val):
    if pd.isna(start_val) or pd.isna(end_val):
        return None

    try:
        start_str = str(start_val).strip().split()[0]
        end_str = str(end_val).strip().split()[0]

        match_start = re.search(r'(\d{1,2})[:\.\-](\d{2})', start_str)
        match_end = re.search(r'(\d{1,2})[:\.\-](\d{2})', end_str)

        if not match_start or not match_end:
            return None

        h1, m1 = map(int, match_start.groups())
        h2, m2 = map(int, match_end.groups())

        now = datetime.now().replace(second=0, microsecond=0)
        start_dt = now.replace(hour=h1, minute=m1)
        end_dt = now.replace(hour=h2, minute=m2)

        is_next_day = False
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
            is_next_day = True

        duration = (end_dt - start_dt).total_seconds() / 3600

        return {
            'start_dt': start_dt,
            'end_dt': end_dt,
            'start_str': f"{h1:02}:{m1:02}",
            'end_str': f"{h2:02}:{m2:02}",
            'duration': round(duration, 2),
            'is_next_day': is_next_day
        }
    except Exception as e:
        return None


def calculate_rest_duration(end_str_N, is_next_day_N, start_str_N1):
    try:
        h_end, m_end = map(int, end_str_N.split(':'))
        h_start, m_start = map(int, start_str_N1.split(':'))
        base_dt = datetime.now().replace(second=0, microsecond=0)
        end_dt_N = base_dt.replace(hour=h_end, minute=m_end)
        if not is_next_day_N:
            end_dt_N -= timedelta(days=1)
        start_dt_N1 = base_dt.replace(hour=h_start, minute=m_start)
        while start_dt_N1 <= end_dt_N:
            start_dt_N1 += timedelta(days=1)
        rest_duration = (start_dt_N1 - end_dt_N).total_seconds() / 3600
        return round(rest_duration, 1)
    except Exception as e:
        return "Н/Д"


def load_history(day_num):
    filename = os.path.join(HISTORY_JSON_DIR, f"history_{day_num}.json")
    if not os.path.exists(filename):
        return {}
    with open(filename, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_history(day_num, data):
    filename = os.path.join(HISTORY_JSON_DIR, f"history_{day_num}.json")
    clean_data = {}
    for k, v in data.items():
        clean_data[k] = {key: val for key, val in v.items() if not isinstance(val, datetime)}

    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(clean_data, f, indent=4, ensure_ascii=False)
    print(f"[Память] Данные о сменах сохранены в {filename}")


def get_schedule_slots(file_path, start_row, step, col_start, col_end, shift_code):
    df = pd.read_excel(file_path, sheet_name="Расписание", header=None)
    slots = []
    current_row = start_row

    print(f"АНАЛИЗ {shift_code} СМЕНЫ (Колонки {col_start}/{col_end})")

    while True:
        pd_idx = current_row - 1
        if pd_idx >= len(df):
            break

        start_val = df.iloc[pd_idx, col_start]
        end_val = df.iloc[pd_idx, col_end]

        time_info = get_shift_info_from_two_cols(start_val, end_val)

        if time_info:
            time_info['shift_code'] = shift_code
            slots.append({
                'excel_row': current_row,
                'time_info': time_info
            })

        current_row += step

    return slots


def get_available_drivers(file_path, day_num, shift_code):
    df = pd.read_excel(file_path, sheet_name="Весь_табель")
    df.columns = df.columns.astype(str)

    day_col = str(day_num)
    if day_col not in df.columns:
        if f'{day_num}.0' in df.columns:
            day_col = f'{day_num}.0'
        else:
            raise ValueError(f"Ошибка: В табеле нет колонки с названием '{day_num}'")

    drivers = []
    for idx, row in df.iterrows():
        tab_no = row.iloc[0]
        if pd.isna(tab_no):
            continue
        try:
            tab_str = str(int(float(tab_no)))
        except (ValueError, TypeError):
            continue

        status_clean = str(row[day_col]).strip()
        if status_clean.startswith(str(shift_code)):
            drivers.append(tab_str)

    return drivers


# --- НОВАЯ ЛОГИКА ВЫБОРА: считаем реальный отдых в часах для водителя относительно целевой смены ---
def get_rest_hours_for_driver(driver_id, history_data, target_shift_start_dt):
    drv_str = str(driver_id)
    if drv_str not in history_data:
        # Нет истории - считаем, что водитель хорошо отдохнул (подходит)
        # Возвращаем значение, равное REST_HOURS, чтобы он считался "идеальным"
        return REST_HOURS

    last_work = history_data[drv_str]
    last_end_str = last_work.get('end_str')
    was_next_day = last_work.get('is_next_day', False)

    try:
        # Определяем дату окончания прошлой смены относительно даты целевой смены
        base_date = target_shift_start_dt.date() - timedelta(days=1)
        if was_next_day:
            base_date = base_date + timedelta(days=1)

        h, m = map(int, last_end_str.split(':'))
        last_end_dt = datetime(base_date.year, base_date.month, base_date.day, h, m)

        rest_hours = (target_shift_start_dt - last_end_dt).total_seconds() / 3600
        # округлим до 1 знака
        return round(rest_hours, 1)
    except Exception:
        return -9999.0  # ошибка - сильно не подходит


# helper: узнаём, работал ли вчера в той же смене (приоритет)
def worked_same_shift_yesterday(driver_id, history_data, shift_code):
    drv_str = str(driver_id)
    if drv_str not in history_data:
        return False
    return history_data[drv_str].get('shift_code') == shift_code


# Новая функция выбора водителя для слота: минимизируем отклонение от REST_HOURS,
# но не допускаем физического наложения смен (т.е. если фактический конец позже старта слота, фильтруем)
def choose_driver_for_slot(candidates, history_data, target_shift_start_dt, shift_code, already_assigned_set):
    scored = []
    for drv in list(candidates):
        if str(drv) in already_assigned_set:
            continue

        rest_h = get_rest_hours_for_driver(drv, history_data, target_shift_start_dt)

        # Если rest_h очень негативный (ошибка расчёта) — пропускаем
        if rest_h == -9999.0:
            continue

        # Если прошлое окончание смены после начала целевой смены -> отрицательное rest_h (пересечение)
        # Это физическое пересечение — не допускаем.
        if rest_h < -0.5:  # почти наверняка пересечение
            continue

        # Наличие истории - для приоритета
        same_shift = 0 if worked_same_shift_yesterday(drv, history_data, shift_code) else 1

        # Если водитель имеет отдыха >= REST_HOURS — это "подходящие" кандидаты.
        # Оценка: чем ближе rest_h к REST_HOURS, тем лучше -> минимизируем abs(...)
        if rest_h >= REST_HOURS:
            score = (abs(rest_h - REST_HOURS), same_shift, -rest_h)  # минимизируем; вторичный критерий - тот, кто работал вчера в той же смене
        else:
            # если недостаточно отдыха — делаем худший приоритет, но всё же можем выбрать, если других нет
            deficit = REST_HOURS - rest_h
            score = (1000 + deficit, same_shift, -rest_h)  # сильно penalize those with less rest

        scored.append((score, drv, rest_h))

    if not scored:
        return None, None

    # выбираем минимальный score
    scored.sort(key=lambda x: x[0])
    chosen = scored[0]
    chosen_driver = chosen[1]
    chosen_rest = chosen[2]
    return str(chosen_driver), chosen_rest


def run_planner():
    print(f"ЗАПУСК ПЛАНИРОВЩИКА: МАРШРУТ №{ROUTE_NUMBER}, ДЕНЬ {TARGET_DAY} (С учетом истории дня {PREV_DAY})\n")

    history = load_history(PREV_DAY)
    today_history_log = {}

    slots_s1 = get_schedule_slots(FILE_PATH, ROW_START, STEP, COL_SHIFT_1_START, COL_SHIFT_1_END, 1)
    slots_s2 = get_schedule_slots(FILE_PATH, ROW_START, STEP, COL_SHIFT_2_START, COL_SHIFT_2_END, 2)

    try:
        candidates_s1 = get_available_drivers(FILE_PATH, TARGET_DAY, 1)
        candidates_s2 = get_available_drivers(FILE_PATH, TARGET_DAY, 2)
    except ValueError as e:
        print(e)
        return

    print(f"[Спрос] 1 смена: {len(slots_s1)} нарядов, 2 смена: {len(slots_s2)} нарядов")
    print(f"[Табель] Доступно: 1 смена = {len(candidates_s1)}, 2 смена = {len(candidates_s2)}")

    wb = load_workbook(FILE_PATH)
    ws = wb["Расписание"]

    assigned_s1 = []
    assigned_s2 = []
    assigned_today_set = set()  # чтобы водитель не получил более одной работы за день

    # --- РАСПРЕДЕЛЕНИЕ 1 СМЕНЫ ---
    for slot in slots_s1:
        slot_start = slot['time_info']['start_dt']

        chosen_driver, chosen_rest = choose_driver_for_slot(candidates_s1, history, slot_start, 1, assigned_today_set)

        if not chosen_driver:
            ws.cell(row=slot['excel_row'], column=COL_SHIFT_1_INSERT).value = "НЕТ_РЕЗЕРВА"
            continue

        ws.cell(row=slot['excel_row'], column=COL_SHIFT_1_INSERT).value = chosen_driver

        # Запишем в историю
        info = slot['time_info'].copy()
        info['shift_code'] = 1
        today_history_log[str(chosen_driver)] = info

        # Удаляем выбранного водителя из обоих списков кандидатов (чтобы не дать 2 смены в день)
        if chosen_driver in candidates_s1:
            candidates_s1.remove(chosen_driver)
        if chosen_driver in candidates_s2:
            candidates_s2.remove(chosen_driver)

        assigned_s1.append(chosen_driver)
        assigned_today_set.add(str(chosen_driver))

    # --- ОБРАБОТКА ОСТАТКОВ 1 СМЕНЫ (РЕЗЕРВ) ---
    for reserve_driver in list(candidates_s1):
        # Если водитель уже назначен (после удаления) — пропускаем
        if str(reserve_driver) in assigned_today_set:
            continue

        now = datetime.now().replace(second=0, microsecond=0)
        dummy_start = now.replace(hour=5, minute=0)
        dummy_end = now.replace(hour=14, minute=0)
        is_next = False
        if dummy_end < dummy_start:
            dummy_end += timedelta(days=1)
            is_next = True

        today_history_log[str(reserve_driver)] = {
            'start_dt': dummy_start,
            'end_dt': dummy_end,
            'start_str': "05:00",
            'end_str': "14:00",
            'duration': round((dummy_end - dummy_start).total_seconds() / 3600, 1),
            'is_next_day': is_next,
            'shift_code': 1,
            'note': 'РЕЗЕРВ'
        }

    # --- РАСПРЕДЕЛЕНИЕ 2 СМЕНЫ ---
    for slot in slots_s2:
        slot_start = slot['time_info']['start_dt']

        chosen_driver, chosen_rest = choose_driver_for_slot(candidates_s2, history, slot_start, 2, assigned_today_set)

        if not chosen_driver:
            ws.cell(row=slot['excel_row'], column=COL_SHIFT_2_INSERT).value = "НЕТ_РЕЗЕРВА"
            continue

        ws.cell(row=slot['excel_row'], column=COL_SHIFT_2_INSERT).value = chosen_driver

        info = slot['time_info'].copy()
        info['shift_code'] = 2
        today_history_log[str(chosen_driver)] = info

        if chosen_driver in candidates_s2:
            candidates_s2.remove(chosen_driver)
        if chosen_driver in candidates_s1:
            candidates_s1.remove(chosen_driver)

        assigned_s2.append(chosen_driver)
        assigned_today_set.add(str(chosen_driver))

    # --- ОБРАБОТКА ОСТАТКОВ 2 СМЕНЫ (РЕЗЕРВ) ---
    for reserve_driver in list(candidates_s2):
        if str(reserve_driver) in assigned_today_set:
            continue

        now = datetime.now().replace(second=0, microsecond=0)
        dummy_start = now.replace(hour=15, minute=0)
        dummy_end = now.replace(hour=23, minute=59)
        is_next = False
        if dummy_end < dummy_start:
            dummy_end += timedelta(days=1)
            is_next = True

        today_history_log[str(reserve_driver)] = {
            'start_dt': dummy_start,
            'end_dt': dummy_end,
            'start_str': "15:00",
            'end_str': "23:59",
            'duration': round((dummy_end - dummy_start).total_seconds() / 3600, 1),
            'is_next_day': is_next,
            'shift_code': 2,
            'note': 'РЕЗЕРВ'
        }

    print(f"\n[Итог] Назначено на маршрут: 1см={len(assigned_s1)}, 2см={len(assigned_s2)}")
    print(f"[Итог] В резерве (без маршрута): 1см={len([d for d in candidates_s1 if str(d) not in assigned_today_set])}, 2см={len([d for d in candidates_s2 if str(d) not in assigned_today_set])}")

    file_path = os.path.join(OUTPUT_DIR, f"Расписание_Итог_{TARGET_DAY}.xlsx")
    wb.save(file_path)
    save_history(TARGET_DAY, today_history_log)


def generate_work_summary(target_day, file_path):
    print("\n\n=== ГЕНЕРАЦИЯ СВОДНОГО ОТЧЕТА ===")

    try:
        df_tab = pd.read_excel(file_path, sheet_name="Весь_табель")  # ← ИЗМЕНЕНО
        driver_id_col = df_tab.columns[0]
        graphik_col = df_tab.columns[1]

        df_temp = df_tab.set_index(df_tab[driver_id_col].astype(str))

        driver_graphik = df_temp[graphik_col].to_dict()
        all_drivers = []
        for x in df_tab[driver_id_col].dropna():
            try:
                all_drivers.append(str(int(float(x))))
            except (ValueError, TypeError):
                continue
    except Exception as e:
        print(f"Ошибка чтения Весь_табель для отчета: {e}")
        return None

    history_data = {}
    for day in range(1, target_day + 2):
        filename = os.path.join(HISTORY_JSON_DIR, f"history_{day}.json")
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                history_data[day] = json.load(f)
        else:
            history_data[day] = {}

    columns = ['График'] + [f'День {d}' for d in range(1, target_day + 1)]
    report_data = []

    for driver_id in all_drivers:
        row = {'График': driver_graphik.get(driver_id, 'N/A')}

        for day in range(1, target_day + 1):
            day_history = history_data.get(day, {})

            if driver_id in day_history:

                work_data = day_history[driver_id]
                duration_N = work_data['duration']
                end_str_N = work_data['end_str']
                is_next_day_N = work_data.get('is_next_day', False)

                next_day_history = history_data.get(day + 1, {})

                if driver_id in next_day_history:
                    start_str_N1 = next_day_history[driver_id]['start_str']
                    rest_h = calculate_rest_duration(end_str_N, is_next_day_N, start_str_N1)
                    rest_report = f"Отдых: {rest_h} ч."
                elif day == target_day:
                    rest_report = "Отдых: Н/Д"
                else:
                    rest_report = f"Отдых: Полный"

                row[f'День {day}'] = f"Работа: {duration_N} ч. | {rest_report}"
            else:
                row[f'День {day}'] = "--- Выходной"

        report_data.append(row)

    df_report = pd.DataFrame(report_data, index=pd.Index(all_drivers, name='Таб. №'))

    report_filename = os.path.join(OUTPUT_DIR, f"Отчет_Нагрузки_Дни_1_по_{target_day}.xlsx")
    df_report.to_excel(report_filename)
    print(f"\n[Отчет] Создан сводный отчет: **{report_filename}**")

    return report_filename


def init_route(route_number):
    """Инициализация путей и файлов для конкретного маршрута"""
    global ROUTE_NUMBER, HISTORY_JSON_DIR, OUTPUT_DIR, FILE_PATH
    
    ROUTE_NUMBER = route_number
    HISTORY_JSON_DIR = f"history_json_route_{ROUTE_NUMBER}"
    OUTPUT_DIR = f"output_route_{ROUTE_NUMBER}"
    
    os.makedirs(HISTORY_JSON_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Пробуем найти файл с номером маршрута, иначе используем data.xlsx
    route_file = f"data_{ROUTE_NUMBER}.xlsx"
    if os.path.exists(route_file):
        FILE_PATH = route_file
    else:
        FILE_PATH = "../data/data.xlsx"
        print(f"[Внимание] Файл {route_file} не найден, используется {FILE_PATH}")
    
    print(f"[Маршрут] Инициализирован маршрут №{ROUTE_NUMBER}")
    print(f"[Файлы] Данные: {FILE_PATH}, История: {HISTORY_JSON_DIR}, Вывод: {OUTPUT_DIR}")


def auto_run_simulation(total_days, file_path, route_number=None):
    """Запуск симуляции для указанного маршрута"""
    if route_number is not None:
        init_route(route_number)
    
    print(f"\n--- ПОДГОТОВКА: Удаление старых log-файлов (history_X.json) ---\n")
    for d in range(1, total_days + 2):
        filename = os.path.join(HISTORY_JSON_DIR, f"history_{d}.json")
        if os.path.exists(filename):
            os.remove(filename)

    for day in range(1, total_days + 1):
        global TARGET_DAY
        global PREV_DAY

        TARGET_DAY = day
        PREV_DAY = day - 1

        print(f"ЗАПУСК ДНЯ {TARGET_DAY:02d}")

        run_planner()

    print("\n")
    print("##################### СИМУЛЯЦИЯ ЗАВЕРШЕНА #####################")
    generate_work_summary(total_days, file_path)


if __name__ == "__main__":
    import sys
    
    # Можно указать номер маршрута как аргумент командной строки
    # Пример: python new_model.py 9
    if len(sys.argv) > 1:
        try:
            route_num = int(sys.argv[1])
            init_route(route_num)
        except ValueError:
            print(f"[Ошибка] Некорректный номер маршрута: {sys.argv[1]}")
            print(f"[Использование] python new_model.py [номер_маршрута]")
            print(f"[По умолчанию] Используется маршрут {ROUTE_NUMBER}")
    
    auto_run_simulation(TOTAL_DAYS_IN_MONTH, FILE_PATH)
