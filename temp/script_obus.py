import openpyxl
import os
import re


def sanitize_filename(name):
    """Очищает имя файла от недопустимых символов"""
    return re.sub(r'[\\/*?:"<>|]', "_", name.strip()) or "unnamed_sheet"


# Создаем папку output, если её нет
os.makedirs("../output_obus", exist_ok=True)

try:
    # Загружаем исходный файл
    workbook = openpyxl.load_workbook("../data/data_obus.xlsx", data_only=False)
except FileNotFoundError:
    print("❌ Ошибка: Файл data_obus.xlsx не найден в текущей директории.")
    exit(1)

# Проверяем количество листов
all_sheets = workbook.sheetnames
if len(all_sheets) < 5:
    print(f"⚠️ В файле всего {len(all_sheets)} листов. Обработка невозможна.")
    exit(0)

print(f"📂 Найдено листов для обработки: {len(all_sheets[4:])}")

# Обрабатываем листы начиная с 5-го (индекс 4)
for sheet_name in all_sheets[2:]:
    source_sheet = workbook[sheet_name]

    # Создаем новую книгу
    new_wb = openpyxl.Workbook()
    # Удаляем стандартный лист "Sheet"
    if "Sheet" in new_wb.sheetnames:
        new_wb.remove(new_wb["Sheet"])

    # Создаем лист с фиксированным названием "Лист1"
    new_sheet = new_wb.create_sheet(title="Лист1")

    # Копируем данные и формулы
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            # Копируем стили и форматирование
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()

    # Формируем безопасное имя файла из оригинального названия листа
    filename = sanitize_filename(sheet_name)
    output_path = os.path.join("../output_obus", f"{filename}.xlsx")

    # Сохраняем результат
    new_wb.save(output_path)
    print(f"✅ Создан файл: {output_path} (внутри: лист 'Лист1')")

print("\n✨ Все листы успешно обработаны! В каждом файле лист назван 'Лист1'.")